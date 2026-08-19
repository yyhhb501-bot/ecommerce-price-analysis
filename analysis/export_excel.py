"""导出项目成果：多 Sheet Excel（统计/价格带/相关性/核心带/定价建议/明细）。"""
import json
import os

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from config import PROCESSED_DIR, BASE_DIR

OUT_XLSX = os.path.join(BASE_DIR, "output", "项目成果分析.xlsx")


def _frames(result):
    summary = pd.DataFrame(result["summary"])
    price_band = pd.DataFrame(result["price_band"])

    corr = pd.DataFrame(
        [{"品类": k, "Pearson": v["pearson"], "Spearman": v["spearman"], "样本数": v["n"]}
         for k, v in result["correlation"].items()]
    )

    core = pd.DataFrame(
        [{"品类": k, "核心价格带": v["core_band"], "销量占比%": v["core_band_sales_share"],
          "平均销量": v["core_band_mean_sales"], "SKU数": v["core_band_sku_count"]}
         for k, v in result["core_price_band"].items()]
    )

    rec = pd.DataFrame(
        [{"品类": k, **v} for k, v in result["recommendations"].items()]
    )

    band_detail = pd.DataFrame(
        [{"品类": k, **d} for k, v in result["core_price_band"].items()
         for d in v["band_detail"]]
    )

    detail = pd.read_csv(os.path.join(PROCESSED_DIR, "products_clean.csv"))
    return summary, price_band, corr, core, rec, band_detail, detail


def _style(workbook):
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    for ws in workbook.worksheets:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for col in ws.columns:
            width = max(len(str(c.value)) for c in col if c.value is not None)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 2, 40)
        ws.freeze_panes = "A2"


def main(result, ab_result=None):
    frames = _frames(result)
    names = ["品类平台统计", "价格带分布", "相关性", "核心价格带",
             "定价建议", "价格带明细", "清洗后明细"]

    if ab_result:
        frames = list(frames) + [
            pd.DataFrame([ab_result["design"]]),
            pd.DataFrame(ab_result["results"]),
            pd.DataFrame(ab_result["group_totals"]),
        ]
        names = names + ["AB实验设计", "AB测试结果", "AB分组汇总"]

    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        for name, df in zip(names, frames):
            df.to_excel(writer, sheet_name=name, index=False)

    from openpyxl import load_workbook
    wb = load_workbook(OUT_XLSX)
    _style(wb)
    wb.save(OUT_XLSX)
    print(f"成果 Excel 已导出: {OUT_XLSX}")
    return OUT_XLSX


if __name__ == "__main__":
    with open(os.path.join(PROCESSED_DIR, "analysis_result.json"), encoding="utf-8") as f:
        main(json.load(f))
